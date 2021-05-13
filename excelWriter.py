# import openpyxl
from openpyxl import workbook  # pip install openpyxl
from openpyxl import load_workbook
from fileDownloader import get_mayafiles_htmlFIle, get_requiredData_from_htmlURL


def writeToExcel(linksArray, pageIndex):
    wb = load_workbook("somepath/workbook.xlsx")
    sheets = wb.sheetnames
    Sheet1 = wb[sheets[0]]
    index_to_add = Sheet1.max_row + 1

    for link in linksArray:
        Sheet1.cell(row=index_to_add, column=1).value = index_to_add
        Sheet1.cell(row=index_to_add, column=2).value = link
        Sheet1.cell(row=index_to_add, column=3).value = pageIndex

        index_to_add += 1

    wb.save("somepath/workbook.xlsx")
    print("done")


def writeToExcel2(excel_path, start, end):
    wb = load_workbook(excel_path)
    sheets = wb.sheetnames
    linksSheet = wb[sheets[0]]
    htmlLinksSheet = wb[sheets[1]]

    # for index in range(2, linksSheet.max_row + 1):
    #     pass

    for index in range(start, end):
        try:
            print("Start: index ", index)
            newLink = get_mayafiles_htmlFIle(
                linksSheet['B' + str(index)].value)
            htmlLinksSheet.cell(row=index, column=1).value = index
            htmlLinksSheet.cell(
                row=index, column=2).value = newLink
            htmlLinksSheet.cell(
                row=index, column=3).value = linksSheet['A' + str(index)].value
            print("Finish: index ", index)

        except:
            htmlLinksSheet.cell(row=index, column=1).value = index
            htmlLinksSheet.cell(row=index, column=2).value = ""
            htmlLinksSheet.cell(row=index, column=3).value = ""

    wb.save("somepath/workbook.xlsx")
    print("done")


def get_subLinks():
    print("beggining first part!!")
    writeToExcel2("somepath/workbook.xlsx", 101, 250)
    print("finish first part")

    print("beggining second part!!")
    writeToExcel2("somepath/workbook.xlsx", 250, 500)
    print("finish second part")

    print("beggining third part!!")
    writeToExcel2("somepath/workbook.xlsx", 500, 750)
    print("finish third part")

    print("beggining 4 part!!")
    writeToExcel2("somepath/workbook.xlsx", 750, 1000)
    print("finish 4 part")

    print("beggining 5 part!!")
    writeToExcel2("somepath/workbook.xlsx", 1000, 1250)
    print("finish 5 part")

    print("beggining 6 part!!")
    writeToExcel2("somepath/workbook.xlsx", 1250, 1500)
    print("finish 6 part")

    print("beggining 7 part!!")
    writeToExcel2("somepath/workbook.xlsx", 1500, 1750)
    print("finish 7 part")

    print("beggining 8 part!!")
    writeToExcel2("somepath/workbook.xlsx", 1750, 2000)
    print("finish 8 part")

    print("beggining 9 part!!")
    writeToExcel2("somepath/workbook.xlsx", 2000, 2201)
    print("finish 9 part")

    print("Finish all parts!!")

# get the missing links from writeToExcel2:


def writeToExcel3(excel_path):
    wb = load_workbook(excel_path)
    sheets = wb.sheetnames
    linksSheet = wb[sheets[0]]
    htmlLinksSheet = wb[sheets[1]]
    count = 0
    for index in range(2, 2201):
        try:
            # print("Start: index ", index)
            newLink = htmlLinksSheet['B' + str(index)].value
            if newLink == None:
                count += 1
                print("index =", index)
                # requiredLink = get_mayafiles_htmlFIle(
                #     linksSheet['B' + str(index)].value)  # get required link
                # htmlLinksSheet.cell(
                #     row=index, column=2).value = requiredLink  # add the link to the sheet

        except:
            pass
            # htmlLinksSheet.cell(row=index, column=1).value = index
            # htmlLinksSheet.cell(row=index, column=2).value = ""
            # htmlLinksSheet.cell(row=index, column=3).value = ""

    print("count = ", count)
    print("done")


def writeDataFromLink(excel_path):
    wb = load_workbook(excel_path)
    sheets = wb.sheetnames
    htmlLinksSheet = wb[sheets[1]]
    DataSheet = wb[sheets[2]]
    index_to_add = DataSheet.max_row + 1
    # initial_parentIndex = DataSheet['G' + str(index_to_add - 1)].value + 1
    initial_parentIndex = 2
    for parentIndex in range(initial_parentIndex, 2201):
        if parentIndex != 1166:
            try:
                print("Start: index ", parentIndex)
                link = htmlLinksSheet['B' + str(parentIndex)].value
                parentId = htmlLinksSheet['A' + str(parentIndex)].value
                # print("start reading index", parentIndex)
                data = get_requiredData_from_htmlURL(link)
                # print("finish reading")
                # add the link to the sheet
                for found in data["foundsData"]:
                    # print(found)
                    DataSheet.cell(row=index_to_add,
                                   column=1).value = index_to_add
                    DataSheet.cell(row=index_to_add,
                                   column=2).value = found["foundId"]
                    DataSheet.cell(row=index_to_add,
                                   column=3).value = found["foundName"]
                    DataSheet.cell(row=index_to_add,
                                   column=4).value = found["valueBefore"]
                    DataSheet.cell(row=index_to_add,
                                   column=5).value = found["valueAfter"]
                    DataSheet.cell(row=index_to_add,
                                   column=6).value = data["sendDate"]
                    DataSheet.cell(row=index_to_add,
                                   column=7).value = data["changeDate"]
                    DataSheet.cell(row=index_to_add, column=8).value = parentId

                    index_to_add += 1

            except Exception as e:
                print("parentIndex - ", parentIndex, " made an exception")
                print(e)

    wb.save("somepath/workbook.xlsx")
    print("done")


def main():
    # linksAraay = ['https://maya.tase.co.il/reports/details/729585', 'https://maya.tase.co.il/reports/details/729241', 'https://maya.tase.co.il/reports/details/729239', 'https://maya.tase.co.il/reports/details/728900', 'https://maya.tase.co.il/reports/details/728831', 'https://maya.tase.co.il/reports/details/728828', 'https://maya.tase.co.il/reports/details/728581', 'https://maya.tase.co.il/reports/details/728320', 'https://maya.tase.co.il/reports/details/728319', 'https://maya.tase.co.il/reports/details/728250', 'https://maya.tase.co.il/reports/details/728012', 'https://maya.tase.co.il/reports/details/728010', 'https://maya.tase.co.il/reports/details/727602', 'https://maya.tase.co.il/reports/details/726968', 'https://maya.tase.co.il/reports/details/725263',
    #               'https://maya.tase.co.il/reports/details/723924', 'https://maya.tase.co.il/reports/details/723020', 'https://maya.tase.co.il/reports/details/722990', 'https://maya.tase.co.il/reports/details/722959', 'https://maya.tase.co.il/reports/details/722891', 'https://maya.tase.co.il/reports/details/722524', 'https://maya.tase.co.il/reports/details/722255', 'https://maya.tase.co.il/reports/details/721922', 'https://maya.tase.co.il/reports/details/721918', 'https://maya.tase.co.il/reports/details/721705', 'https://maya.tase.co.il/reports/details/721168', 'https://maya.tase.co.il/reports/details/721167', 'https://maya.tase.co.il/reports/details/721166', 'https://maya.tase.co.il/reports/details/721163', 'https://maya.tase.co.il/reports/details/721162']
    # writeToExcel(linksAraay)
    # get_subLinks()
    writeDataFromLink("somepath/workbook.xlsx")


if __name__ == "__main__":
    main()
